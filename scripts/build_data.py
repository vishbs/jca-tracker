"""
Build data.json for the JCA tracker site from the EU HTA Excel export.

Source: https://health.ec.europa.eu/document/download/d947533e-7e4e-4e82-a9c6-e06830d708f8_en?filename=hta_ongoing-jca_en.xlsx
Manual update flow: replace data/hta_ongoing-jca_en.xlsx with a freshly
downloaded copy, then re-run this script to regenerate src/data/jcas.json.
"""

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "hta_ongoing-jca_en.xlsx"
OUT_PATH = ROOT / "src" / "data" / "jcas.json"
META_OUT_PATH = ROOT / "src" / "data" / "meta.json"

HEADER_ROW = 14  # 0-indexed for pandas -> row 15 in Excel
EXTRACTED_ON_CELL = "A11"  # each sheet has its own "Data extracted on <date>" note here
SHEETS = {
    "Ongoing JCAs": "Ongoing",
    "Discontinued JCAs": "Discontinued",
    "Completed JCAs": "Completed",
}

# Only a few columns carry document links in each sheet.
LINK_COLUMNS = {
    "Links to relevant documents": "links_relevant_documents",
    "Link to JCA report and related documents": "link_jca_report",
    "Link to product information on the EMA website": "link_ema_product_info",
    "Link to the Union Register of medicinal products for human use": "link_union_register",
}


def slugify(text):
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def get_country(agency_string):
    if pd.isna(agency_string):
        return None
    return agency_string.split(",")[-1].strip()


def fmt_date(value):
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def extract_data_extracted_on():
    """Read the 'Data extracted on <date>' note from each sheet and return the latest."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    dates = []
    for sheet_name in SHEETS:
        text = wb[sheet_name][EXTRACTED_ON_CELL].value
        if not isinstance(text, str):
            continue
        match = re.search(r"Data extracted on\s+(.+)", text)
        if not match:
            continue
        try:
            dates.append(datetime.strptime(match.group(1).strip(), "%d %B %Y").date())
        except ValueError:
            continue
    return max(dates) if dates else None


def extract_hyperlinks(sheet_name, columns):
    """Map (excel_row, column_name) -> hyperlink URL for the given sheet."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[sheet_name]
    header_cells = {
        cell.value: cell.column
        for cell in ws[HEADER_ROW + 1]
        if cell.value in columns
    }
    links = {}
    for col_name, col_idx in header_cells.items():
        for row in ws.iter_rows(min_row=HEADER_ROW + 2, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.hyperlink:
                links[(cell.row, col_name)] = cell.hyperlink.target
    return links


def load_sheet(sheet_name, status):
    df = pd.read_excel(XLSX_PATH, sheet_name=sheet_name, header=HEADER_ROW)
    df = df.dropna(how="all").reset_index(drop=True)

    present_link_cols = [c for c in LINK_COLUMNS if c in df.columns]
    hyperlinks = extract_hyperlinks(sheet_name, present_link_cols) if present_link_cols else {}

    records = []
    for i, row in df.iterrows():
        excel_row = HEADER_ROW + 2 + i  # +1 for header itself, +1 for 1-indexing

        name = row.get("Name of medicine")
        inn = row.get("International non-proprietary name (INN) / Common Name")
        display_name = name if isinstance(name, str) and name.strip() else inn

        record = {
            "status": status,
            "medicine_name": display_name if isinstance(display_name, str) else None,
            "inn": inn if isinstance(inn, str) else None,
            "indication": row.get("Indication - Summary") if isinstance(row.get("Indication - Summary"), str) else None,
            "substance_type": row.get("Substance type (classification)") if isinstance(row.get("Substance type (classification)"), str) else None,
            "assessor": row.get("Assessor") if isinstance(row.get("Assessor"), str) else None,
            "assessor_country": get_country(row.get("Assessor")),
            "coassessor": row.get("Co-assessor") if isinstance(row.get("Co-assessor"), str) else None,
            "coassessor_country": get_country(row.get("Co-assessor")),
            "orphan_product": row.get("Orphan product") if isinstance(row.get("Orphan product"), str) else None,
            "accelerated_assessment": row.get("Accelerated Assessment (Art. 14(9) Reg  726/2004)") if isinstance(row.get("Accelerated Assessment (Art. 14(9) Reg  726/2004)"), str) else None,
            "revert_to_standard_timetable": row.get("Revert to standard Time Table (MM/YY)") if isinstance(row.get("Revert to standard Time Table (MM/YY)"), str) else None,
            "variation_to_existing_ma": row.get("Variation to the terms of an existing MA") if isinstance(row.get("Variation to the terms of an existing MA"), str) else None,
            "date_ema_validation": fmt_date(row.get("Date of EMA validation of the MAA")),
            "date_jca_discontinuation": fmt_date(row.get("Date of JCA discontinuation")),
            "reason_for_discontinuation": row.get("Reason for discontinuation") if isinstance(row.get("Reason for discontinuation"), str) else None,
            "date_marketing_authorisation": fmt_date(row.get("Date of marketing authorisation")),
            "date_htacg_endorsement": fmt_date(row.get("Date of HTACG endorsement of the JCA report")),
            "date_ec_procedural_review": fmt_date(row.get("Date of conclusion of the procedural review by the European Commission")),
            "date_report_publication": fmt_date(row.get("Date of publication of the JCA report")),
        }

        for excel_col, json_key in LINK_COLUMNS.items():
            record[json_key] = hyperlinks.get((excel_row, excel_col))

        slug_source = record["medicine_name"] or record["inn"] or f"{status}-{i}"
        record["slug"] = slugify(f"{slug_source}-{status}")

        records.append(record)

    return records


def main():
    all_records = []
    for sheet_name, status in SHEETS.items():
        all_records.extend(load_sheet(sheet_name, status))

    all_records.sort(key=lambda r: (r["status"] != "Ongoing", r["medicine_name"] or ""))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w") as f:
        json.dump(all_records, f, indent=2)

    data_extracted_on = extract_data_extracted_on()
    with open(META_OUT_PATH, "w") as f:
        json.dump(
            {"data_extracted_on": data_extracted_on.isoformat() if data_extracted_on else None},
            f,
            indent=2,
        )

    counts = {}
    for r in all_records:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    print(f"Wrote {len(all_records)} JCAs to {OUT_PATH}")
    print(f"  {counts}")
    print(f"Data extracted on: {data_extracted_on} -> {META_OUT_PATH}")


if __name__ == "__main__":
    main()
